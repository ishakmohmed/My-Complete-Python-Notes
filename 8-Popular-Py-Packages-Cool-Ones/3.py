import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("NewSheet", 0)
# 0 is index
# wb.remove_sheet()
# pass a sheet object to remove_sheet()

cell = sheet["a2"]
# a2 is coordinate of cell
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)


sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

columns = sheet["a"]
print(columns)

columns2 = sheet["a:c"]
print(columns2)

columns3 = sheet["a1:c3"]
print(columns3)

rows = sheet[1:3]  # will return rows 1 till 3 cause argument(s) = number

sheet.append([1, 2, 3])  # to add a row at end

# please be aware of insert_rows, insert_cols, delete_rows, and delete_cols too

wb.save("transactions2.xlsx")
